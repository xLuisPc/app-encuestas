from rest_framework import viewsets, status, generics
from rest_framework.decorators import action, api_view, permission_classes, authentication_classes
from rest_framework.response import Response as DRFResponse
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.views import APIView
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import json
from django.db.models import Count, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
from .models import Survey, Question, Response, Answer, Option, MatrixRow, MatrixColumn
from .serializers import (
    SurveySerializer, SurveyPublicSerializer, QuestionSerializer,
    ResponseSerializer
)
from .permissions import (
    IsAdminOrCreator, IsSurveyCreatorOrAdmin, 
    IsAssignedViewerOrAdmin, CanViewStatistics
)


class SurveyViewSet(viewsets.ModelViewSet):
    serializer_class = SurveySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.is_admin():
            return Survey.objects.all()
        elif user.is_creator():
            return Survey.objects.filter(creator=user)
        else:  # viewer
            return Survey.objects.filter(
                Q(assigned_viewers=user) | Q(creator=user)
            ).distinct()

    def get_permissions(self):
        if self.action == 'list' or self.action == 'retrieve':
            return [IsAuthenticated()]
        elif self.action == 'create':
            return [IsAdminOrCreator()]
        elif self.action in ['update', 'partial_update', 'destroy']:
            return [IsSurveyCreatorOrAdmin()]
        return super().get_permissions()

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    @action(detail=True, methods=['get'], permission_classes=[CanViewStatistics])
    def statistics(self, request, pk=None):
        """Obtener estadísticas de una encuesta"""
        try:
            survey = self.get_object()
        except Survey.DoesNotExist:
            return DRFResponse(
                {'error': 'Encuesta no encontrada.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            stats = {
                'survey': {
                    'id': str(survey.id),
                    'title': survey.title,
                    'total_responses': survey.total_responses,
                },
                'questions': []
            }
            
            for question in survey.questions.all():
                question_stats = {
                    'id': question.id,
                    'text': question.text,
                    'question_type': question.question_type,  # Cambiar 'type' a 'question_type' para consistencia
                    'total_answers': 0,
                    'data': {}
                }
                
                try:
                    if question.question_type == 'single':
                        # Contar respuestas por opción
                        answers = Answer.objects.filter(
                            question=question,
                            selected_option__isnull=False
                        ).values('selected_option__text').annotate(
                            count=Count('id')
                        )
                        question_stats['data'] = {
                            option['selected_option__text']: option['count']
                            for option in answers
                        }
                        question_stats['total_answers'] = sum(question_stats['data'].values()) if question_stats['data'] else 0 if question_stats['data'] else 0
                    
                    elif question.question_type == 'multiple':
                        # Contar respuestas por opción (puede haber múltiples por respuesta)
                        answers = Answer.objects.filter(
                            question=question,
                            selected_option__isnull=False
                        ).values('selected_option__text').annotate(
                            count=Count('id')
                        )
                        question_stats['data'] = {
                            option['selected_option__text']: option['count']
                            for option in answers
                        }
                        question_stats['total_answers'] = Answer.objects.filter(
                            question=question
                        ).values('response').distinct().count()
                    
                    elif question.question_type == 'matrix':
                        # Contar respuestas por fila y columna
                        matrix_data = {}
                        answers = Answer.objects.filter(
                            question=question,
                            matrix_row__isnull=False,
                            matrix_column__isnull=False
                        ).values('matrix_row__text', 'matrix_column__text').annotate(
                            count=Count('id')
                        )
                        
                        for answer in answers:
                            row = answer.get('matrix_row__text', '')
                            col = answer.get('matrix_column__text', '')
                            if row and col:
                                if row not in matrix_data:
                                    matrix_data[row] = {}
                                matrix_data[row][col] = answer['count']
                        
                        question_stats['data'] = matrix_data
                        question_stats['total_answers'] = Answer.objects.filter(
                            question=question
                        ).values('response').distinct().count()
                except Exception as e:
                    # Si hay error procesando una pregunta, continuar con las demás
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f'Error procesando pregunta {question.id}: {str(e)}')
                    question_stats['data'] = {}
                    question_stats['total_answers'] = 0
                
                stats['questions'].append(question_stats)
            
            return DRFResponse(stats)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error en statistics: {str(e)}', exc_info=True)
            return DRFResponse(
                {'error': f'Error al generar estadísticas: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], permission_classes=[CanViewStatistics])
    def export_excel(self, request, pk=None):
        """Exportar estadísticas a Excel con gráficas (una pregunta por página)"""
        survey = self.get_object()
        
        wb = Workbook()
        # Eliminar la hoja por defecto
        wb.remove(wb.active)
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        title_font = Font(bold=True, size=16)
        question_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=12)
        
        question_num = 1
        for question in survey.questions.all():
            # Crear una hoja nueva para cada pregunta
            ws = wb.create_sheet(title=f"Pregunta {question_num}")
            
            # Encabezado de la encuesta
            ws['A1'] = f"Encuesta: {survey.title}"
            ws['A1'].font = title_font
            ws.merge_cells('A1:F1')
            
            ws['A2'] = f"Total de respuestas: {survey.total_responses}"
            ws['A2'].font = subtitle_font
            ws.merge_cells('A2:F2')
            
            row = 4
            
            # Título de pregunta
            ws[f'A{row}'] = f"Pregunta {question_num}: {question.text}"
            ws[f'A{row}'].font = question_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 2
            
            if question.question_type in ['single', 'multiple']:
                # Encabezados de tabla
                ws[f'A{row}'] = "Opción"
                ws[f'B{row}'] = "Cantidad"
                ws[f'C{row}'] = "Porcentaje"
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{row}']
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                # Calcular total de respuestas (distintas por respuesta para multiple)
                if question.question_type == 'single':
                    total = Answer.objects.filter(
                        question=question,
                        selected_option__isnull=False
                    ).count()
                else:  # multiple
                    total = Answer.objects.filter(
                        question=question
                    ).values('response').distinct().count()
                
                # Obtener datos
                answers = Answer.objects.filter(
                    question=question,
                    selected_option__isnull=False
                ).values('selected_option__text').annotate(
                    count=Count('id')
                ).order_by('-count')
                
                data_start_row = row
                for answer in answers:
                    option_text = answer['selected_option__text']
                    count = answer['count']
                    percentage = (count / total * 100) if total > 0 else 0
                    
                    ws[f'A{row}'] = option_text
                    ws[f'B{row}'] = count
                    ws[f'C{row}'] = f"{percentage:.1f}%"
                    # Alinear datos
                    ws[f'A{row}'].alignment = Alignment(horizontal='left')
                    ws[f'B{row}'].alignment = Alignment(horizontal='center')
                    ws[f'C{row}'].alignment = Alignment(horizontal='center')
                    row += 1
                
                data_end_row = row - 1
                
                # Espaciado antes de las gráficas
                row += 2
                
                # Gráficas (igual que en la web: pastel y barras)
                if total > 0 and data_end_row >= data_start_row:
                    # Gráfica de pastel (superior, columna E)
                    pie_chart = PieChart()
                    pie_chart.title = f"Pregunta {question_num}"
                    data_ref = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row)
                    cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
                    pie_chart.add_data(data_ref, titles_from_data=False)
                    pie_chart.set_categories(cats_ref)
                    pie_chart.width = 14
                    pie_chart.height = 8
                    pie_chart.legend.position = 'r'  # Right
                    ws.add_chart(pie_chart, f'E{row}')
                    
                    # Gráfica de barras (inferior, columna E, más abajo)
                    bar_chart = BarChart()
                    bar_chart.type = "col"
                    bar_chart.style = 10
                    bar_chart.title = f"Pregunta {question_num}"
                    bar_chart.y_axis.title = 'Cantidad'
                    bar_chart.x_axis.title = 'Opción'
                    bar_chart.add_data(data_ref, titles_from_data=False)
                    bar_chart.set_categories(cats_ref)
                    bar_chart.width = 14
                    bar_chart.height = 8
                    bar_chart.legend.position = 'r'
                    # Colocar la gráfica de barras más abajo
                    chart_row = row + 18
                    ws.add_chart(bar_chart, f'E{chart_row}')
            
            elif question.question_type == 'matrix':
                # Obtener columnas y filas ordenadas
                columns = question.matrix_columns.all().order_by('order')
                rows_data = question.matrix_rows.all().order_by('order')
                col_headers = [col.text for col in columns]
                
                # Encabezados de tabla
                ws[f'A{row}'] = "Fila / Columna"
                ws[f'A{row}'].fill = header_fill
                ws[f'A{row}'].font = header_font
                ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                for idx, col_header in enumerate(col_headers, start=1):
                    cell = ws.cell(row=row, column=idx+1)
                    cell.value = col_header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
                
                # Datos por fila
                data_start_row = row
                
                for matrix_row in rows_data:
                    ws[f'A{row}'] = matrix_row.text
                    ws[f'A{row}'].alignment = Alignment(horizontal='left')
                    for idx, col in enumerate(columns, start=1):
                        count = Answer.objects.filter(
                            question=question,
                            matrix_row=matrix_row,
                            matrix_column=col
                        ).count()
                        cell = ws.cell(row=row, column=idx+1, value=count)
                        cell.alignment = Alignment(horizontal='center')
                    row += 1
                
                data_end_row = row - 1
                
                # Espaciado antes de las gráficas
                row += 2
                
                # Para matrices, dejar espacio para gráfica de pastel (aunque no se muestre)
                if data_end_row >= data_start_row and len(col_headers) > 0:
                    # Espacio reservado para gráfica de pastel (no se crea, solo espacio)
                    
                    # Gráfica de barras agrupadas (igual que en la web)
                    bar_chart = BarChart()
                    bar_chart.type = "col"
                    bar_chart.style = 10
                    bar_chart.title = f"Pregunta {question_num}"
                    bar_chart.y_axis.title = 'Cantidad'
                    bar_chart.x_axis.title = 'Opciones'
                    
                    # Referencia de datos: incluir encabezados de columnas
                    data_ref = Reference(ws, min_col=2, min_row=data_start_row-1, 
                                       max_col=len(col_headers)+1, max_row=data_end_row)
                    cats_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=data_end_row)
                    bar_chart.add_data(data_ref, titles_from_data=True)
                    bar_chart.set_categories(cats_ref)
                    bar_chart.width = 14
                    bar_chart.height = 8
                    bar_chart.legend.position = 'b'  # Bottom
                    # Colocar la gráfica de barras más abajo para dejar espacio arriba
                    chart_row = row + 18
                    ws.add_chart(bar_chart, f'E{chart_row}')
            
            # Ajustar ancho de columnas para esta hoja
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 2
            ws.column_dimensions['E'].width = 2
            ws.column_dimensions['F'].width = 2
            
            # Congelar encabezados
            ws.freeze_panes = 'A4'
            
            question_num += 1
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="encuesta_{survey.id}.xlsx"'
        wb.save(response)
        return response


class SurveyPublicView(generics.RetrieveAPIView):
    """Vista pública para ver y responder encuestas"""
    queryset = Survey.objects.prefetch_related(
        'questions__options',
        'questions__matrix_rows',
        'questions__matrix_columns'
    ).all()
    serializer_class = SurveyPublicSerializer
    permission_classes = [AllowAny]
    authentication_classes = []  # No requiere autenticación
    lookup_field = 'id'

    def get(self, request, *args, **kwargs):
        try:
            survey = self.get_object()
        except Survey.DoesNotExist:
            return DRFResponse(
                {'error': 'Encuesta no encontrada.'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if not survey.is_active:
            return DRFResponse(
                {'error': 'Esta encuesta no está activa.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not survey.is_open:
            # Formatear fechas en UTC para el mensaje (el frontend las convertirá a hora local)
            from django.utils import timezone
            start_str = survey.start_date.isoformat() if hasattr(survey.start_date, 'isoformat') else str(survey.start_date)
            end_str = survey.end_date.isoformat() if hasattr(survey.end_date, 'isoformat') else str(survey.end_date)
            return DRFResponse(
                {'error': f'Esta encuesta no está disponible actualmente. Fechas: {start_str} - {end_str}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().get(request, *args, **kwargs)


@csrf_exempt
def ResponseCreateView(request):
    """Vista para crear respuestas a encuestas (público - sin autenticación)"""
    from django.contrib.auth.models import AnonymousUser
    from django.utils import timezone
    from rest_framework.request import Request
    from rest_framework.parsers import JSONParser
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)
    
    # Parsear JSON manualmente
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON inválido'}, status=400)
    
    survey_id = data.get('survey')
    try:
        survey = Survey.objects.get(id=survey_id)
    except Survey.DoesNotExist:
        return JsonResponse(
            {'error': 'Encuesta no encontrada.'},
            status=404
        )
    
    # Verificar que la encuesta esté activa y dentro de las fechas
    now = timezone.now()
    
    if not survey.is_active:
        return JsonResponse(
            {'error': 'Esta encuesta no está activa.'},
            status=400
        )
    
        if now < survey.start_date:
            start_str = survey.start_date.isoformat() if hasattr(survey.start_date, 'isoformat') else str(survey.start_date)
            return JsonResponse(
                {'error': f'Esta encuesta aún no está disponible. Inicia el {start_str}.'},
                status=400
            )
        
        if now > survey.end_date:
            end_str = survey.end_date.isoformat() if hasattr(survey.end_date, 'isoformat') else str(survey.end_date)
            return JsonResponse(
                {'error': f'Esta encuesta ya cerró el {end_str}.'},
                status=400
            )
    
    try:
        # Log para debug
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f'Datos recibidos: survey={data.get("survey")}, answers count={len(data.get("answers", []))}')
        logger.info(f'Datos completos: {json.dumps(data, default=str)}')
        
        serializer = ResponseSerializer(data=data, context={'request': request})
        if not serializer.is_valid():
            logger.error(f'Errores de validación: {serializer.errors}')
            return JsonResponse(
                {'error': f'Datos inválidos: {serializer.errors}'},
                status=400
            )
        
        response_obj = serializer.save()
        logger.info(f'Respuesta guardada: {response_obj.id}, respuestas creadas: {response_obj.answers.count()}')
        
        return JsonResponse(
            {'message': 'Respuesta guardada exitosamente.'},
            status=201
        )
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f'Error en ResponseCreateView: {str(e)}')
        logger.error(traceback.format_exc())
        return JsonResponse(
            {'error': f'Error al guardar la respuesta: {str(e)}'},
            status=400
        )

